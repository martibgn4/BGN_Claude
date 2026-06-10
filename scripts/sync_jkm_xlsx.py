"""Sync the JKM master.csv into a multi-sheet Excel workbook.

Sheets:
    1. Settlements   — long-format raw data (settlement_date, strip_label, month_iso, settlement, key)
    2. Curve Matrix  — wide pivot: rows = settlement_date desc, cols = month_iso, values = settlement
    3. Lookup        — two dropdowns (date, contract) and a cell that returns the settlement
    4. Metadata      — generated timestamp, row counts, date range, source path

Usage:
    python scripts/sync_jkm_xlsx.py
        # writes to data/reports/JKM_settlements.xlsx by default

    python scripts/sync_jkm_xlsx.py --out "C:/Users/me/OneDrive/Desk/JKM_settlements.xlsx"
        # writes to a shared location (OneDrive / SharePoint sync / network drive)

    set JKM_XLSX_OUT=C:\\Users\\me\\OneDrive\\Desk\\JKM_settlements.xlsx
    python scripts/sync_jkm_xlsx.py
        # env var override; useful for setting once in Windows User env vars

Requires openpyxl: `python -m pip install openpyxl`.
"""
import argparse
import csv
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path


def _import_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        return Workbook, Alignment, Font, PatternFill, get_column_letter, DataValidation
    except ImportError as e:
        raise ImportError(
            "openpyxl not installed. Install via:\n"
            "  python -m pip install openpyxl"
        ) from e


def read_master(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _contract_sort_key(c: str) -> tuple:
    """Sort YYYY-MM ascending, then 'bal:*' entries after them."""
    return (1, c) if c.startswith("bal:") else (0, c)


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------
def _settlements_sheet(ws, rows, styles):
    Font_, PatternFill_, Alignment_, gcl = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    header_font = Font_(bold=True, color="FFFFFF")
    header_fill = PatternFill_("solid", fgColor="1F4E78")

    # The key column lets the Lookup sheet do a plain MATCH (no array formula needed)
    columns = [
        ("settlement_date",   14),
        ("strip_label",       18),
        ("month_iso",         14),
        ("settlement",        14),
        ("source_screenshot", 22),
        ("key",               28),
    ]
    for i, (h, w) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill = header_font, header_fill
        ws.column_dimensions[gcl(i)].width = w

    for r, row in enumerate(rows, start=2):
        d  = row.get("settlement_date", "")
        sl = row.get("strip_label", "")
        mi = row.get("month_iso", "")
        ws.cell(row=r, column=1, value=d)
        ws.cell(row=r, column=2, value=sl)
        ws.cell(row=r, column=3, value=mi)
        try:
            ws.cell(row=r, column=4, value=float(row.get("settlement", ""))).number_format = "#,##0.000"
        except (TypeError, ValueError):
            ws.cell(row=r, column=4, value=row.get("settlement", ""))
        ws.cell(row=r, column=5, value=row.get("source_screenshot", ""))
        ws.cell(row=r, column=6, value=f"{d}|{mi}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _matrix_sheet(ws, rows, styles):
    Font_, PatternFill_, _, gcl = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    header_font = Font_(bold=True, color="FFFFFF")
    header_fill = PatternFill_("solid", fgColor="1F4E78")

    matrix: dict[str, dict[str, float]] = defaultdict(dict)
    contracts: set[str] = set()
    dates: set[str] = set()
    for r in rows:
        d, c = r.get("settlement_date", ""), r.get("month_iso", "")
        try:
            v = float(r.get("settlement", ""))
        except (TypeError, ValueError):
            continue
        if not d or not c:
            continue
        matrix[d][c] = v
        contracts.add(c)
        dates.add(d)

    contracts_sorted = sorted(contracts, key=_contract_sort_key)
    dates_sorted     = sorted(dates, reverse=True)  # most recent first

    h = ws.cell(row=1, column=1, value="settlement_date \\ contract")
    h.font, h.fill = header_font, header_fill
    for j, c in enumerate(contracts_sorted, start=2):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font, cell.fill = header_font, header_fill

    for i, d in enumerate(dates_sorted, start=2):
        ws.cell(row=i, column=1, value=d).font = Font_(bold=True)
        for j, c in enumerate(contracts_sorted, start=2):
            v = matrix[d].get(c)
            if v is not None:
                cell = ws.cell(row=i, column=j, value=v)
                cell.number_format = "#,##0.000"

    ws.column_dimensions["A"].width = 16
    for j in range(2, len(contracts_sorted) + 2):
        ws.column_dimensions[gcl(j)].width = 10
    ws.freeze_panes = "B2"


def _lookup_sheet(ws, rows, styles):
    Font_, PatternFill_, _, _ = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    DataValidation_ = styles["DataValidation"]

    dates_sorted     = sorted({r["settlement_date"] for r in rows}, reverse=True)
    contracts_sorted = sorted({r["month_iso"] for r in rows}, key=_contract_sort_key)
    non_bal_first    = next((c for c in contracts_sorted if not c.startswith("bal:")), "")

    ws.cell(row=1, column=1, value="JKM Settlement Lookup").font = Font_(bold=True, size=14, color="1F4E78")
    ws.cell(row=2, column=1, value="Pick a settlement date and a contract; the result cell pulls from Settlements via INDEX/MATCH.").font = Font_(italic=True, color="666666")

    ws.cell(row=4, column=1, value="Settlement date:").font = Font_(bold=True)
    ws.cell(row=5, column=1, value="Contract (month_iso):").font = Font_(bold=True)
    ws.cell(row=7, column=1, value="Settlement ($/MMBtu):").font = Font_(bold=True)

    # Default selections — most recent date, first real (non-bal) contract
    if dates_sorted:
        ws.cell(row=4, column=2, value=dates_sorted[0])
    if non_bal_first:
        ws.cell(row=5, column=2, value=non_bal_first)
    elif contracts_sorted:
        ws.cell(row=5, column=2, value=contracts_sorted[0])

    # Input cells styling
    input_fill = PatternFill_("solid", fgColor="FFF2CC")
    ws.cell(row=4, column=2).fill = input_fill
    ws.cell(row=5, column=2).fill = input_fill

    # Result formula — plain (non-array) MATCH against the 'key' column we built
    # in the Settlements sheet (column F). Works in any modern Excel version.
    result_formula = (
        '=IFERROR(INDEX(Settlements!D:D, '
        'MATCH(B4 & "|" & B5, Settlements!F:F, 0)), "Not found")'
    )
    res = ws.cell(row=7, column=2, value=result_formula)
    res.number_format = "#,##0.000"
    res.font = Font_(bold=True, size=14, color="1F4E78")

    # Hidden helper columns G (dates) and H (contracts) — feed the dropdowns
    ws.cell(row=1, column=7, value="Dates")
    for i, d in enumerate(dates_sorted, start=2):
        ws.cell(row=i, column=7, value=d)
    ws.cell(row=1, column=8, value="Contracts")
    for i, c in enumerate(contracts_sorted, start=2):
        ws.cell(row=i, column=8, value=c)
    ws.column_dimensions["G"].hidden = True
    ws.column_dimensions["H"].hidden = True

    if dates_sorted:
        dv_date = DataValidation_(
            type="list",
            formula1=f"=$G$2:$G${len(dates_sorted) + 1}",
            allow_blank=False,
        )
        dv_date.add("B4")
        ws.add_data_validation(dv_date)
    if contracts_sorted:
        dv_contract = DataValidation_(
            type="list",
            formula1=f"=$H$2:$H${len(contracts_sorted) + 1}",
            allow_blank=False,
        )
        dv_contract.add("B5")
        ws.add_data_validation(dv_contract)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22


def _metadata_sheet(ws, rows, master_path: Path, out_path: Path, styles):
    Font_, _, _, _ = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    dates     = sorted({r["settlement_date"] for r in rows})
    contracts = sorted({r["month_iso"] for r in rows}, key=_contract_sort_key)

    ws.cell(row=1, column=1, value="JKM Settlements — Master Workbook").font = Font_(bold=True, size=14)
    items = [
        ("Generated (local time)", datetime.now().isoformat(timespec="seconds")),
        ("Source file",            str(master_path)),
        ("Output file",            str(out_path)),
        ("Total rows",             len(rows)),
        ("Settlement dates",       len(dates)),
        ("Earliest date",          dates[0] if dates else ""),
        ("Latest date",            dates[-1] if dates else ""),
        ("Unique contracts",       len(contracts)),
        ("Front contract",         next((c for c in contracts if not c.startswith("bal:")), "")),
        ("Back contract",          contracts[-1] if contracts else ""),
    ]
    for i, (k, v) in enumerate(items, start=3):
        ws.cell(row=i, column=1, value=k).font = Font_(bold=True)
        ws.cell(row=i, column=2, value=v)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def write_workbook(rows, master_path: Path, out_path: Path) -> None:
    Workbook, Alignment, Font, PatternFill, get_column_letter, DataValidation = _import_openpyxl()
    styles = {
        "Alignment": Alignment, "Font": Font, "PatternFill": PatternFill,
        "get_column_letter": get_column_letter, "DataValidation": DataValidation,
    }
    wb = Workbook()
    wb.remove(wb.active)
    _settlements_sheet(wb.create_sheet("Settlements"),    rows, styles)
    _matrix_sheet     (wb.create_sheet("Curve Matrix"),   rows, styles)
    _lookup_sheet     (wb.create_sheet("Lookup"),         rows, styles)
    _metadata_sheet   (wb.create_sheet("Metadata"),       rows, master_path, out_path, styles)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Sync JKM master.csv into an Excel workbook.")
    ap.add_argument("--master", default="data/raw/jkm_settlements/master.csv",
                    help="Path to master CSV (relative to repo root or absolute).")
    ap.add_argument("--out", default=None,
                    help="Output xlsx path. If omitted, uses JKM_XLSX_OUT env var, "
                         "else falls back to data/reports/JKM_settlements.xlsx (local).")
    args = ap.parse_args()

    repo_root = Path(__file__).resolve().parent.parent

    master_path = Path(args.master)
    if not master_path.is_absolute():
        master_path = repo_root / master_path

    out_arg = args.out or os.environ.get("JKM_XLSX_OUT") or "data/reports/JKM_settlements.xlsx"
    out_path = Path(out_arg)
    if not out_path.is_absolute():
        out_path = repo_root / out_path

    if not master_path.exists():
        print(f"ERROR: master file not found: {master_path}", file=sys.stderr)
        return 2

    rows = read_master(master_path)
    if not rows:
        print(f"ERROR: master file is empty: {master_path}", file=sys.stderr)
        return 2

    try:
        write_workbook(rows, master_path, out_path)
        # write_workbook(rows, master_path, Path("C:\\Users\\marti.fernandezreal\\BAYEGAN DIS TIC. A.S\\LNG Team - 01. Miscellaneous\\14. BGN Tools"))

    except ImportError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2
    except OSError as e:
        # Common when shared path isn't reachable (OneDrive offline, network share down, file open in Excel)
        print(f"ERROR writing to {out_path}: {e}", file=sys.stderr)
        print("Hint: check the shared location is mounted, the file isn't open in Excel, "
              "and JKM_XLSX_OUT (if set) points somewhere reachable.", file=sys.stderr)
        return 1

    dates = sorted({r["settlement_date"] for r in rows})
    contracts = sorted({r["month_iso"] for r in rows}, key=_contract_sort_key)
    print(f"Wrote {out_path}")
    print(f"  Rows: {len(rows)}")
    print(f"  Settlement dates: {len(dates)} ({dates[0]} -> {dates[-1]})")
    print(f"  Contracts: {len(contracts)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
