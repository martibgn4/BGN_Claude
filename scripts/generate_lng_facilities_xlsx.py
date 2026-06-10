"""Generate the consolidated LNG facilities Excel workbook from the CSV database.

Reads from `data/reference/lng_facilities/{liquefaction,regasification,disruptions}.csv`
and writes a multi-sheet workbook to `data/reports/lng_facilities.xlsx`.

Sheets:
    1. Summary
    2. Liquefaction - Operational
    3. Liquefaction - Upcoming
    4. Regasification - Operational
    5. Regasification - Upcoming
    6. Capacity Schedule (year-by-year liquefaction additions)
    7. Disruptions
    8. Sources

Requires openpyxl. Install via:
    python -m pip install openpyxl
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable


def _import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        return openpyxl, Alignment, Font, PatternFill, get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl not installed. Install via:\n"
            "  python -m pip install openpyxl"
        ) from e


# ---------------------------------------------------------------------------
# CSV reading
# ---------------------------------------------------------------------------
def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _to_float(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_year(d: str) -> int | None:
    if not d:
        return None
    try:
        return datetime.strptime(d.strip(), "%Y-%m-%d").year
    except ValueError:
        try:
            return int(d.strip()[:4])
        except (TypeError, ValueError):
            return None


# ---------------------------------------------------------------------------
# Status classification
# ---------------------------------------------------------------------------
OPERATIONAL = {"operational"}
UPCOMING    = {"planned", "FID", "construction", "commissioning"}
INACTIVE    = {"mothballed", "retired", "paused"}


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
LIQ_COLUMNS = [
    ("facility_id",     "Facility ID",     14),
    ("country",         "Country",         12),
    ("region",          "Region",          16),
    ("facility_name",   "Facility",        32),
    ("operator",        "Operator",        32),
    ("train",           "Train",           10),
    ("status",          "Status",          14),
    ("nameplate_mtpa",  "Nameplate (MTPA)", 14),
    ("effective_mtpa",  "Effective (MTPA)", 14),
    ("fid_date",        "FID",             12),
    ("start_date",      "Start / COD",     12),
    ("ramp_up",         "Ramp-up",         24),
    ("feedgas",         "Feedgas",         28),
    ("offtake_summary", "Offtake",         44),
    ("source",          "Source",          28),
    ("last_updated",    "Updated",         11),
    ("notes",           "Notes",           38),
]

REGAS_COLUMNS = [
    ("facility_id",     "Facility ID",      14),
    ("country",         "Country",          16),
    ("region",          "Region",           14),
    ("facility_name",   "Facility",         32),
    ("operator",        "Operator",         32),
    ("terminal_type",   "Type",             10),
    ("status",          "Status",           14),
    ("sendout_bcm_y",   "Send-out (Bcm/y)", 14),
    ("sendout_mtpa",    "Send-out (MTPA)",  14),
    ("storage_m3",      "Storage (m³ LNG)", 14),
    ("start_date",      "Start",            12),
    ("reload_capable",  "Reload?",          9),
    ("source",          "Source",           24),
    ("last_updated",    "Updated",          11),
    ("notes",           "Notes",            36),
]

DISRUPTION_COLUMNS = [
    ("facility_id",          "Facility ID",          22),
    ("disruption_id",        "Disruption ID",        24),
    ("disruption_type",      "Type",                 16),
    ("start_date",           "Start",                12),
    ("end_date",             "End",                  12),
    ("capacity_impact_mtpa", "Impact (MTPA)",        14),
    ("description",          "Description",          80),
    ("source",               "Source",               28),
    ("last_updated",         "Updated",              11),
]


def _write_table(ws, rows: list[dict], columns: list[tuple[str, str, int]],
                 styles) -> None:
    Font, PatternFill, Alignment, get_column_letter = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Header
    for i, (_key, label, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data
    numeric_keys = {"nameplate_mtpa", "effective_mtpa", "sendout_bcm_y",
                    "sendout_mtpa", "storage_m3", "capacity_impact_mtpa"}
    for r, row in enumerate(rows, start=2):
        for i, (key, _label, _w) in enumerate(columns, start=1):
            v = row.get(key, "")
            if key in numeric_keys:
                v = _to_float(v) if v != "" else None
            c = ws.cell(row=r, column=i, value=v)
            if key in numeric_keys and v is not None:
                c.number_format = "#,##0.0" if key != "storage_m3" else "#,##0"
            c.alignment = wrap

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _build_summary_sheet(ws, liquefaction: list[dict], regas: list[dict],
                         disruptions: list[dict], styles) -> None:
    Font, PatternFill, Alignment, _ = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    section_fill = PatternFill("solid", fgColor="D9E1F2")

    ws["A1"] = "LNG Facilities — Consolidated Summary"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
    ws["A2"].font = Font(italic=True, color="666666")

    op_liq = [r for r in liquefaction if r["status"] in OPERATIONAL]
    up_liq = [r for r in liquefaction if r["status"] in UPCOMING]
    in_liq = [r for r in liquefaction if r["status"] in INACTIVE]

    op_re  = [r for r in regas if r["status"] in OPERATIONAL]
    up_re  = [r for r in regas if r["status"] in UPCOMING]

    active_disrupt = [d for d in disruptions if (d.get("end_date") or "").strip() == ""]

    def _sum(rows, key):
        return sum(_to_float(r.get(key, 0)) for r in rows)

    # Overall counts
    row = 4
    ws.cell(row=row, column=1, value="Overall").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1
    overall = [
        ("Liquefaction facilities tracked",          len(liquefaction)),
        ("  Operational trains",                     len(op_liq)),
        ("  Upcoming (planned/FID/construction)",    len(up_liq)),
        ("  Mothballed / paused / retired",          len(in_liq)),
        ("Regasification terminals tracked",         len(regas)),
        ("  Operational",                            len(op_re)),
        ("  Upcoming",                               len(up_re)),
        ("Active disruptions",                       len(active_disrupt)),
    ]
    for label, val in overall:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = "#,##0"
        row += 1

    # Capacity totals
    row += 1
    ws.cell(row=row, column=1, value="Capacity (MTPA)").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1
    cap = [
        ("Operational liquefaction nameplate (MTPA)", _sum(op_liq, "nameplate_mtpa")),
        ("Operational liquefaction effective (MTPA)", _sum(op_liq, "effective_mtpa")),
        ("Upcoming liquefaction nameplate (MTPA)",    _sum(up_liq, "nameplate_mtpa")),
        ("Operational regas send-out (MTPA equiv)",   _sum(op_re, "sendout_mtpa")),
        ("Upcoming regas send-out (MTPA equiv)",      _sum(up_re, "sendout_mtpa")),
    ]
    for label, val in cap:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = "#,##0.0"
        row += 1

    # By region (operational liquefaction)
    row += 1
    ws.cell(row=row, column=1, value="Operational Liquefaction by Region").font = header_font
    ws.cell(row=row, column=1).fill = section_fill
    row += 1
    ws.cell(row=row, column=1, value="Region").font = header_font
    ws.cell(row=row, column=2, value="Nameplate (MTPA)").font = header_font
    ws.cell(row=row, column=3, value="Effective (MTPA)").font = header_font
    ws.cell(row=row, column=4, value="Trains").font = header_font
    row += 1
    by_region: dict[str, list[dict]] = defaultdict(list)
    for r in op_liq:
        by_region[r["region"]].append(r)
    for region in sorted(by_region.keys()):
        rs = by_region[region]
        ws.cell(row=row, column=1, value=region)
        ws.cell(row=row, column=2, value=_sum(rs, "nameplate_mtpa")).number_format = "#,##0.0"
        ws.cell(row=row, column=3, value=_sum(rs, "effective_mtpa")).number_format = "#,##0.0"
        ws.cell(row=row, column=4, value=len(rs)).number_format = "#,##0"
        row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10


def _build_schedule_sheet(ws, liquefaction: list[dict], styles) -> None:
    """Year-by-year incremental and cumulative liquefaction additions."""
    Font, PatternFill, _, _ = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    # Aggregate MTPA added per year (operational + upcoming, by start_date year)
    by_year: dict[int, dict[str, float]] = defaultdict(lambda: {"op": 0.0, "up": 0.0})
    for r in liquefaction:
        y = _parse_year(r.get("start_date", ""))
        if y is None:
            continue
        bucket = "op" if r["status"] in OPERATIONAL else ("up" if r["status"] in UPCOMING else None)
        if bucket is None:
            continue
        by_year[y][bucket] += _to_float(r.get("nameplate_mtpa"))

    years = sorted(by_year.keys())
    if not years:
        ws["A1"] = "No dated facilities to schedule."
        return

    # Header
    headers = ["Year", "Additions — operational record (MTPA)",
               "Additions — upcoming (MTPA)", "Total additions (MTPA)",
               "Cumulative since first year (MTPA)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill

    cumulative = 0.0
    for r, y in enumerate(years, start=2):
        op = by_year[y]["op"]
        up = by_year[y]["up"]
        total = op + up
        cumulative += total
        ws.cell(row=r, column=1, value=y)
        ws.cell(row=r, column=2, value=op).number_format = "#,##0.0"
        ws.cell(row=r, column=3, value=up).number_format = "#,##0.0"
        ws.cell(row=r, column=4, value=total).number_format = "#,##0.0"
        ws.cell(row=r, column=5, value=cumulative).number_format = "#,##0.0"

    for letter, width in zip("ABCDE", [10, 36, 24, 22, 32]):
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"


def _build_sources_sheet(ws, liquefaction, regas, disruptions, styles) -> None:
    Font, PatternFill, _, _ = (
        styles["Font"], styles["PatternFill"], styles["Alignment"], styles["get_column_letter"],
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    counts: dict[str, int] = defaultdict(int)
    for rows in (liquefaction, regas, disruptions):
        for r in rows:
            s = (r.get("source") or "").strip()
            if s:
                counts[s] += 1

    for i, h in enumerate(["Source", "Rows referencing"], start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
    for i, (src, n) in enumerate(sorted(counts.items(), key=lambda kv: -kv[1]), start=2):
        ws.cell(row=i, column=1, value=src)
        ws.cell(row=i, column=2, value=n)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def generate(db_dir: Path, out_path: Path) -> tuple[int, int, int]:
    openpyxl, Alignment, Font, PatternFill, get_column_letter = _import_openpyxl()
    styles = {"Alignment": Alignment, "Font": Font, "PatternFill": PatternFill,
              "get_column_letter": get_column_letter}

    liquefaction = _read_csv(db_dir / "liquefaction.csv")
    regas        = _read_csv(db_dir / "regasification.csv")
    disruptions  = _read_csv(db_dir / "disruptions.csv")

    op_liq = [r for r in liquefaction if r["status"] in OPERATIONAL]
    up_liq = [r for r in liquefaction if r["status"] in (UPCOMING | INACTIVE)]
    op_re  = [r for r in regas if r["status"] in OPERATIONAL]
    up_re  = [r for r in regas if r["status"] in UPCOMING]

    op_liq = sorted(op_liq, key=lambda r: (r["region"], r["country"], r["facility_name"], r["train"]))
    up_liq = sorted(up_liq, key=lambda r: (r.get("start_date", ""), r["region"], r["facility_name"]))
    op_re  = sorted(op_re,  key=lambda r: (r["region"], r["country"], r["facility_name"]))
    up_re  = sorted(up_re,  key=lambda r: (r.get("start_date", ""), r["region"], r["facility_name"]))
    disruptions = sorted(disruptions, key=lambda r: (r.get("start_date", ""), r.get("facility_id", "")))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_summary_sheet(wb.create_sheet("Summary"), liquefaction, regas, disruptions, styles)
    _write_table(wb.create_sheet("Liquefaction — Operational"), op_liq, LIQ_COLUMNS, styles)
    _write_table(wb.create_sheet("Liquefaction — Upcoming"),    up_liq, LIQ_COLUMNS, styles)
    _write_table(wb.create_sheet("Regasification — Operational"), op_re, REGAS_COLUMNS, styles)
    _write_table(wb.create_sheet("Regasification — Upcoming"),    up_re, REGAS_COLUMNS, styles)
    _build_schedule_sheet(wb.create_sheet("Capacity Schedule"), liquefaction, styles)
    _write_table(wb.create_sheet("Disruptions"), disruptions, DISRUPTION_COLUMNS, styles)
    _build_sources_sheet(wb.create_sheet("Sources"), liquefaction, regas, disruptions, styles)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return len(liquefaction), len(regas), len(disruptions)


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate the LNG facilities Excel workbook.")
    ap.add_argument("--db-dir", default="data/reference/lng_facilities",
                    help="Folder containing liquefaction.csv, regasification.csv, disruptions.csv")
    ap.add_argument("--out", default="data/reports/lng_facilities.xlsx",
                    help="Output xlsx path.")
    args = ap.parse_args()

    repo_root = Path(__file__).resolve().parent.parent
    db_dir = repo_root / args.db_dir
    out = repo_root / args.out

    n_liq, n_re, n_dis = generate(db_dir, out)
    print(f"Wrote {out}")
    print(f"  Liquefaction rows: {n_liq}")
    print(f"  Regasification rows: {n_re}")
    print(f"  Disruption rows: {n_dis}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
